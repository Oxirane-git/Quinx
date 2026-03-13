#!/usr/bin/env python3
"""
Quinx AI — Master Pipeline Orchestrator

Runs the full lead generation + email writing pipeline in one command:
  Step 1: Email_Scrap  — Google Maps search + website scraping + CSV build
  Step 2: Bridge       — Convert CSV to Email_Writer chunk format
  Step 3: Email_Writer — AI-powered email generation
  Step 4: Pause        — Print instructions to review then send

Usage:
    python run_pipeline.py --niche "bakeries"
    python run_pipeline.py --niche "cafes" --cities "London UK" "Tokyo Japan"
    python run_pipeline.py --niche "restaurants" --cities "New York City USA" --chunk-size 10
"""

import argparse
import csv
import glob
import imaplib
import os
import random
import re
import smtplib
import ssl
import subprocess
import sys
import time
from pathlib import Path
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EMAIL_SCRAP_DIR = BASE_DIR / "Email_Scrap"
EMAIL_WRITER_DIR = BASE_DIR / "Email_Writer"
EMAIL_SENDER_DIR = BASE_DIR / "Email_Sender"

# ---------------------------------------------------------------------------
# Defaults
# ---------------------------------------------------------------------------
HIGH_ROI_CITIES = [
    "New York City USA",
    "Los Angeles USA",
    "Chicago USA",
    "Toronto Canada",
    "Miami USA",
    "Houston USA",
    "Vancouver Canada",
    "Atlanta USA",
    "London UK",
    "Berlin Germany",
    "Amsterdam Netherlands",
    "Paris France",
    "Barcelona Spain",
    "Dubai UAE",
    "Manchester UK",
    "Sydney Australia",
    "Melbourne Australia",
    "Singapore",
    "Tokyo Japan",
    "Mumbai India",
    "Auckland New Zealand",
    "Sao Paulo Brazil",
    "Mexico City Mexico",
    "Buenos Aires Argentina",
    "Bogota Colombia",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(text: str) -> str:
    """Convert text to lowercase hyphenated slug."""
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def run(cmd: list[str], cwd: Path, step: str) -> None:
    """Run a subprocess and abort the pipeline on failure."""
    result = subprocess.run(cmd, cwd=str(cwd))
    if result.returncode != 0:
        print(f"\n[ERROR] Step '{step}' failed (exit code {result.returncode}).", file=sys.stderr)
        print(f"        Command: {' '.join(cmd)}", file=sys.stderr)
        print(f"        Working dir: {cwd}", file=sys.stderr)
        sys.exit(result.returncode)


def newest_file(pattern: str) -> Path | None:
    """Return the most recently modified file matching a glob pattern."""
    matches = glob.glob(pattern)
    if not matches:
        return None
    return Path(max(matches, key=os.path.getmtime))


def banner(msg: str) -> None:
    width = max(len(msg) + 4, 50)
    print("\n" + "=" * width)
    print(f"  {msg}")
    print("=" * width)


# ---------------------------------------------------------------------------
# Step 1: Email_Scrap
# ---------------------------------------------------------------------------

def run_email_scrap(niche: str, cities: list[str], fresh: bool = False) -> Path:
    """Run Maps search + scraping + CSV build. Returns path to output CSV."""
    banner(f"STEP 1 — Scraping leads for: {niche} ({len(cities)} cities)")

    niche_slug = slugify(niche)

    for i, city in enumerate(cities, 1):
        city_slug = slugify(city)
        raw_file = EMAIL_SCRAP_DIR / ".tmp" / f"raw_places_{niche_slug}_{city_slug}.json"
        enriched_file = EMAIL_SCRAP_DIR / ".tmp" / f"enriched_{niche_slug}_{city_slug}.json"

        # Skip if already enriched (resume support)
        if enriched_file.exists() and enriched_file.stat().st_size > 10:
            print(f"  [{i}/{len(cities)}] Skipping {city} — already enriched.")
            continue

        print(f"\n  [{i}/{len(cities)}] {city}")

        # Maps search
        if not (raw_file.exists() and raw_file.stat().st_size > 10):
            run(
                ["python", "tools/google_maps_search.py", "--niche", niche, "--city", city],
                cwd=EMAIL_SCRAP_DIR,
                step=f"google_maps_search ({city})",
            )
        else:
            print(f"    Maps search: cached ({raw_file.name})")

        # Website scraping
        run(
            ["python", "tools/scrape_website_emails.py", "--input", str(raw_file)],
            cwd=EMAIL_SCRAP_DIR,
            step=f"scrape_website_emails ({city})",
        )

    # Build final CSV
    print("\n  Building leads CSV...")
    build_cmd = ["python", "tools/build_leads_csv.py", "--niche", niche]
    if fresh:
        # Pass a non-existent dir so no existing emails are excluded
        build_cmd += ["--exclude-leads-dir", "__none__"]
    run(build_cmd, cwd=EMAIL_SCRAP_DIR, step="build_leads_csv")

    # Find the CSV just created (build_leads_csv.py writes to .tmp/)
    csv_pattern = str(EMAIL_SCRAP_DIR / ".tmp" / f"leads_{niche_slug}_*.csv")
    csv_path = newest_file(csv_pattern)
    if not csv_path:
        print(f"[ERROR] No CSV found matching: {csv_pattern}", file=sys.stderr)
        sys.exit(1)

    print(f"\n  Leads CSV: {csv_path}")
    return csv_path


# ---------------------------------------------------------------------------
# Step 2: Bridge — CSV → Email_Writer chunks
# ---------------------------------------------------------------------------

def convert_to_chunks(csv_path: Path, niche: str, chunk_size: int,
                      max_leads: int | None = None) -> int:
    """
    Read Email_Scrap CSV, apply column mapping, filter to rows with emails,
    apply owner_name fallback, and write chunk XLSX files to Email_Writer/leads/.

    Returns the number of leads written.
    """
    banner("STEP 2 — Converting leads to Email_Writer format")

    leads_dir = EMAIL_WRITER_DIR / "leads"
    leads_dir.mkdir(exist_ok=True)

    # Clear existing chunk files
    for old in leads_dir.glob("leads_chunk_*.xlsx"):
        old.unlink()

    # Read CSV
    rows = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = (row.get("email") or "").strip()
            if not email:
                continue  # Email_Writer requires an email address

            business_name = (row.get("business_name") or "").strip()
            raw_owner = (row.get("owner_name") or "").strip()
            owner_name = raw_owner if raw_owner else f"{business_name}'s Team"

            rows.append({
                "business_name": business_name,
                "email": email,
                "phone": (row.get("phone") or "").strip(),
                "website": (row.get("website") or "").strip(),
                "city": (row.get("city") or "").strip(),
                "category": (row.get("category") or "").strip(),
                "niche": niche,
                "owner_name": owner_name,
            })

    if not rows:
        print("[ERROR] No leads with email addresses found in the CSV.", file=sys.stderr)
        sys.exit(1)

    if max_leads is not None:
        rows = rows[:max_leads]
        print(f"  Capped to {max_leads} leads (--max-leads)")

    # Write chunks
    columns = ["business_name", "email", "phone", "website", "city", "category", "niche", "owner_name"]
    total_chunks = (len(rows) + chunk_size - 1) // chunk_size

    for chunk_idx in range(total_chunks):
        chunk_rows = rows[chunk_idx * chunk_size : (chunk_idx + 1) * chunk_size]
        chunk_num = chunk_idx + 1
        chunk_file = leads_dir / f"leads_chunk_{chunk_num:03d}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(columns)
        for lead in chunk_rows:
            ws.append([lead[col] for col in columns])
        wb.save(chunk_file)

    print(f"  {len(rows)} leads -> {total_chunks} chunk(s) written to {leads_dir}")
    return len(rows)


# ---------------------------------------------------------------------------
# Step 3: Email_Writer
# ---------------------------------------------------------------------------

def run_email_writer() -> Path:
    """Run batch_write_emails.py and return the output Excel path."""
    banner("STEP 3 — Generating personalized emails (AI)")

    run(
        ["python", "tools/batch_write_emails.py"],
        cwd=EMAIL_WRITER_DIR,
        step="batch_write_emails",
    )

    output_pattern = str(EMAIL_WRITER_DIR / "emails" / "email_output*.xlsx")
    output_path = newest_file(output_pattern)
    if not output_path:
        print("[ERROR] Email_Writer produced no output file.", file=sys.stderr)
        sys.exit(1)

    return output_path


# ---------------------------------------------------------------------------
# Step 4: Send emails (SMTP + IMAP save-to-Sent)
# ---------------------------------------------------------------------------

SMTP_HOST  = "smtp.hostinger.com"
SMTP_PORT  = 465
IMAP_HOST  = "imap.hostinger.com"
IMAP_PORT  = 993
FROM_EMAIL = "team@tryquinx.com"
FROM_NAME  = "Sahil | Quinx AI"

def _load_smtp_password() -> str:
    env_file = EMAIL_SENDER_DIR / ".env"
    if env_file.exists():
        for line in env_file.read_text().splitlines():
            if line.startswith("HOSTINGER_PASSWORD="):
                return line.split("=", 1)[1].strip()
    return ""


def _find_sent_folder(imap: imaplib.IMAP4_SSL) -> str:
    """Return the correct Sent folder name for this Hostinger account."""
    candidates = ["Sent", "INBOX.Sent", "Sent Items", "Sent Messages"]
    _, folders = imap.list()
    folder_names = []
    for f in folders or []:
        parts = f.decode().split('"."')
        if parts:
            folder_names.append(parts[-1].strip().strip('"'))
    for candidate in candidates:
        if candidate in folder_names:
            return candidate
    # Fall back to first candidate that doesn't error
    for candidate in candidates:
        try:
            status, _ = imap.select(candidate)
            if status == "OK":
                imap.close()
                return candidate
        except Exception:
            pass
    return "Sent"


def send_emails(output_path: Path) -> None:
    """Send all leads from output Excel via SMTP and save each to IMAP Sent."""
    banner("STEP 4 — Sending emails (SMTP + saving to Sent folder)")

    password = _load_smtp_password()
    if not password:
        print("[ERROR] HOSTINGER_PASSWORD not found in Email_Sender/.env", file=sys.stderr)
        sys.exit(1)

    # Load leads from Excel
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("email") and d.get("subject") and d.get("body"):
            leads.append(d)

    if not leads:
        print("[ERROR] No sendable leads found in output file.", file=sys.stderr)
        sys.exit(1)

    print(f"  Loaded {len(leads)} leads from {output_path.name}")

    ctx = ssl.create_default_context()
    sent_count, failed_count = 0, 0

    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as smtp:
        smtp.login(FROM_EMAIL, password)

        with imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT) as imap:
            imap.login(FROM_EMAIL, password)
            sent_folder = _find_sent_folder(imap)
            imap.select(sent_folder)
            print(f"  IMAP Sent folder: {sent_folder}")

            for i, lead in enumerate(leads, 1):
                to_email = lead["email"]
                subject  = lead["subject"]
                body     = lead["body"]
                biz      = lead["business_name"]

                # Build RFC-compliant message
                msg = MIMEMultipart("alternative")
                msg["From"]    = f"{FROM_NAME} <{FROM_EMAIL}>"
                msg["To"]      = to_email
                msg["Subject"] = subject
                msg["Date"]    = time.strftime("%a, %d %b %Y %H:%M:%S +0000", time.gmtime())
                msg.attach(MIMEText(body, "plain", "utf-8"))
                raw = msg.as_bytes()

                try:
                    smtp.sendmail(FROM_EMAIL, to_email, raw)
                    # Save copy to Sent folder via IMAP
                    imap.append(
                        sent_folder,
                        "\\Seen",
                        imaplib.Time2Internaldate(time.time()),
                        raw,
                    )
                    sent_count += 1
                    ts = time.strftime("%H:%M:%S")
                    print(f"  [{ts}] [{i}/{len(leads)}] SENT   -> {to_email}  ({biz})")
                except Exception as e:
                    failed_count += 1
                    ts = time.strftime("%H:%M:%S")
                    print(f"  [{ts}] [{i}/{len(leads)}] FAILED -> {to_email}  ({biz}) | {e}")

                if i < len(leads):
                    delay = random.randint(10, 15)
                    print(f"    Waiting {delay}s...")
                    time.sleep(delay)

    print()
    print(f"  Done.  Sent: {sent_count}   Failed: {failed_count}")
    print(f"  Check '{sent_folder}' in Hostinger webmail to confirm.")


# ---------------------------------------------------------------------------
# Step 4 (no-send path): Finish — instruct user to review then send
# ---------------------------------------------------------------------------

def print_finish(leads_count: int, output_path: Path) -> None:
    rel_output = output_path.relative_to(BASE_DIR)
    sender_env = EMAIL_SENDER_DIR / ".env"

    print("\n")
    print("=" * 60)
    print("  PIPELINE COMPLETE")
    print("=" * 60)
    print(f"  Leads processed : {leads_count}")
    print(f"  Output file     : {rel_output}")
    print()
    print("  Next steps:")
    print(f"  1. Open and review: {output_path}")
    print(f"  2. Update LEADS_FILE in: {sender_env}")
    print(f"     LEADS_FILE={output_path}")
    print()
    print("  3. Then send:")
    print(f"     cd {EMAIL_SENDER_DIR}")
    print("     npm start")
    print("=" * 60)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Quinx AI — End-to-end lead scraping + email writing pipeline"
    )
    parser.add_argument(
        "--niche",
        required="--send-only" not in sys.argv,
        help='Business type to target (e.g. "cafes", "bakeries", "restaurants")',
    )
    parser.add_argument(
        "--cities",
        nargs="+",
        default=None,
        metavar="CITY",
        help='Cities to scrape (e.g. "London UK" "Tokyo Japan"). Defaults to 25 high-ROI cities.',
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=20,
        dest="chunk_size",
        help="Number of leads per Email_Writer chunk file (default: 20)",
    )
    parser.add_argument(
        "--max-leads",
        type=int,
        default=None,
        dest="max_leads",
        help="Cap the number of leads passed to Email_Writer (e.g. 10 for a trial run)",
    )
    parser.add_argument(
        "--fresh",
        action="store_true",
        help="Skip dedup against existing Leads/ folder (useful for trial runs)",
    )
    parser.add_argument(
        "--send",
        action="store_true",
        help="Automatically send emails after generation (saves to IMAP Sent folder)",
    )
    parser.add_argument(
        "--send-only",
        metavar="XLSX",
        dest="send_only",
        default=None,
        help="Skip scraping/writing — just send from an existing output Excel file",
    )
    args = parser.parse_args()

    # --send-only: jump straight to sending
    if args.send_only:
        send_emails(Path(args.send_only))
        return

    cities = args.cities if args.cities else HIGH_ROI_CITIES

    print(f"\nQuinx AI Pipeline")
    print(f"  Niche      : {args.niche}")
    print(f"  Cities     : {len(cities)}")
    print(f"  Chunk size : {args.chunk_size}")
    if args.max_leads:
        print(f"  Max leads  : {args.max_leads}  (trial mode)")
    if args.fresh:
        print(f"  Mode       : --fresh (skipping dedup against existing Leads/)")
    if args.send:
        print(f"  Auto-send  : ON (will send after generation)")

    csv_path = run_email_scrap(args.niche, cities, fresh=args.fresh)
    leads_count = convert_to_chunks(csv_path, args.niche, args.chunk_size, args.max_leads)
    output_path = run_email_writer()

    if args.send:
        send_emails(output_path)
    else:
        print_finish(leads_count, output_path)


if __name__ == "__main__":
    main()
