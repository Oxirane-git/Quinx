#!/usr/bin/env python3
"""
Quinx AI — Batch Email Writer

Reads all leads_chunk_*.xlsx files from the leads/ folder, generates
personalized emails via OpenRouter (with key rotation), and writes
all results to a single output Excel file.

Stops gracefully if all API keys are exhausted (token limits hit).

Usage:
    python tools/batch_write_emails.py
    python tools/batch_write_emails.py --output emails/email_output_10chunks.xlsx
"""

import glob
import json
import os
import re
import sys
import time
import traceback

import openpyxl
import requests
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Setup
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

from tools.write_email import (
    OPENROUTER_API_URL,
    MODEL,
    TEMPERATURE,
    MIN_WORDS,
    MAX_WORDS,
    DEFAULT_CAMPAIGN_CONTEXT,
    DEFAULT_SIGN_OFF,
    load_api_keys,
    load_anthropic_key,
    load_master_prompt,
    substitute_placeholders,
    strip_markdown_fences,
    count_words,
    _call_openrouter_single,
    _is_rate_limit_error,
    _call_anthropic_single,
    _is_anthropic_rate_limit,
)
from tools.enrich_lead import enrich_lead

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
LEADS_DIR = os.path.join(BASE_DIR, "leads")
DEFAULT_OUTPUT = os.path.join(BASE_DIR, "emails", "email_output_10chunks.xlsx")
DELAY_BETWEEN_CALLS = 2  # seconds between API calls to avoid rate limits

# Output columns
OUTPUT_COLUMNS = [
    "business_name",
    "email",
    "phone",
    "website",
    "city",
    "category",
    "niche",
    "subject",
    "body",
    "status",
    "failure_reason",
    "chunk_file",
]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def read_all_leads(input_path: str = None) -> list[dict]:
    """Read leads from a specific file or all leads_chunk_*.xlsx files."""
    if input_path:
        if not os.path.exists(input_path):
            print(f"Input file not found: {input_path}", file=sys.stderr)
            sys.exit(1)
        files = [input_path]
    else:
        pattern = os.path.join(LEADS_DIR, "leads_chunk_*.xlsx")
        files = sorted(glob.glob(pattern))
        if not files:
            print(f"No leads_chunk_*.xlsx files found in {LEADS_DIR}", file=sys.stderr)
            sys.exit(1)

    all_leads = []
    for filepath in files:
        chunk_name = os.path.basename(filepath)
        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            lead = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    lead[header] = row[i] if row[i] is not None else ""
                else:
                    lead[header] = ""
            lead["_chunk_file"] = chunk_name
            all_leads.append(lead)

        wb.close()

    return all_leads


def load_campaign_config(campaign: str | None) -> dict:
    """Load campaign config from a JSON file. Returns empty dict if not specified."""
    if not campaign:
        return {}
    if os.path.isabs(campaign) or campaign.endswith(".json"):
        path = campaign
    else:
        campaigns_dir = os.path.join(BASE_DIR, "campaigns")
        path = os.path.join(campaigns_dir, f"{campaign}.json")
    if not os.path.exists(path):
        print(f"Campaign config not found: {path} — service placeholders will be 'Unknown'.", file=sys.stderr)
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_context(lead: dict, api_keys: list[str], key_status: dict, campaign_config: dict | None = None) -> tuple[dict, bool]:
    """Build a businessContext dict, enriched with scraped website data."""
    enrichment, all_exhausted = enrich_lead(lead, api_keys, key_status)
    business_name = lead.get("business_name", "")
    raw_owner = (lead.get("owner_name") or "").strip()
    owner_name = raw_owner if raw_owner else f"{business_name}'s Team"
    context = {
        "businessName": business_name,
        "ownerName": owner_name,
        "city": lead.get("city", ""),
        "category": lead.get("category", ""),
        "website": lead.get("website", "Unknown"),
        "websiteSummary": enrichment.get("websiteSummary", "Unknown"),
        "rating": enrichment.get("rating", "Unknown"),
        "reviewCount": enrichment.get("reviewCount", "Unknown"),
        "positiveThemes": enrichment.get("positiveThemes", "Unknown"),
        "churnSignals": enrichment.get("churnSignals", "Unknown"),
        "lastReviewDate": enrichment.get("lastReviewDate", "Unknown"),
        "socialPresence": enrichment.get("socialPresence", "Unknown"),
        "hasLoyaltyProgram": enrichment.get("hasLoyaltyProgram", "Unknown"),
        "hasEmailCapture": enrichment.get("hasEmailCapture", "Unknown"),
        "recentMentions": enrichment.get("recentMentions", "Unknown"),
        "painScore": enrichment.get("painScore", "6"),
    }
    # Merge campaign config (serviceName, serviceContext, senderName, etc.)
    if campaign_config:
        for key, value in campaign_config.items():
            if key not in context:  # never overwrite lead-specific fields
                context[key] = value
    return context, all_exhausted


def validate_email_output(result: dict, context: dict) -> str | None:
    """Validate the email output. Returns error reason or None if valid."""
    if "subject" not in result or "body" not in result:
        return "Response JSON missing 'subject' or 'body' key."

    subject = result["subject"]
    body = result["body"]

    subject_words = count_words(subject)
    if subject_words > 8:
        return f"Subject line is {subject_words} words. Must be under 9."

    body_words = count_words(body)
    if body_words < MIN_WORDS or body_words > MAX_WORDS:
        return f"Body is {body_words} words. Must be between {MIN_WORDS} and {MAX_WORDS}."

    biz_name = context.get("businessName", "")
    if biz_name and biz_name.lower() not in body.lower():
        return f"Business name '{biz_name}' does not appear in email body."

    return None


def _parse_anthropic_response(response: requests.Response, key_label: str) -> dict | None:
    """Parse a successful Anthropic API response into a result dict. Returns None on failure."""
    try:
        data = response.json()
    except json.JSONDecodeError:
        print(f"  {key_label}: Invalid JSON response.", file=sys.stderr)
        return None

    raw_text = data.get("content", [{}])[0].get("text", "")
    if not raw_text:
        print(f"  {key_label}: Empty response content.", file=sys.stderr)
        return None

    cleaned = strip_markdown_fences(raw_text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        json_match = re.search(r"\{[^{}]*\}", cleaned, re.DOTALL)
        if json_match:
            try:
                return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass
        print(f"  {key_label}: No JSON found in response.", file=sys.stderr)
        return None


def call_api_with_rotation(
    prompt: str,
    api_keys: list[str],
    key_status: dict,
    retry_reason: str | None = None,
    anthropic_key: str | None = None,
) -> tuple[dict | None, str | None, bool]:
    """
    Call OpenRouter API with key rotation. Falls back to Anthropic if all
    OpenRouter keys are exhausted and anthropic_key is provided.

    Returns: (result_dict, error_string, all_keys_exhausted)
    - result_dict: parsed JSON if successful, None on failure
    - error_string: error description if failed
    - all_keys_exhausted: True if ALL keys (incl. Anthropic) exhausted
    """
    full_prompt = prompt
    if retry_reason:
        full_prompt = (
            f"IMPORTANT CORRECTION: Your previous attempt failed because: "
            f"{retry_reason}\n"
            f"Please follow the rules EXACTLY this time. Pay special attention "
            f"to the rule you violated.\n\n"
            f"{prompt}"
        )

    exhausted_count = 0

    for key_index, api_key in enumerate(api_keys, start=1):
        key_label = f"Key #{key_index}"

        # Skip keys that are already known to be exhausted
        if key_status.get(key_index, {}).get("exhausted", False):
            exhausted_count += 1
            continue

        print(f"  {key_label}: Trying...", file=sys.stderr)

        try:
            response = _call_openrouter_single(full_prompt, api_key)
        except requests.exceptions.Timeout:
            print(f"  {key_label}: Timeout. Retrying after 5s...", file=sys.stderr)
            time.sleep(5)
            try:
                response = _call_openrouter_single(full_prompt, api_key)
            except Exception as e2:
                print(f"  {key_label}: Failed after retry: {e2}", file=sys.stderr)
                continue
        except Exception as e:
            print(f"  {key_label}: Request error: {e}", file=sys.stderr)
            continue

        # Rate limited — mark key exhausted and try next
        if _is_rate_limit_error(response):
            try:
                err_body = response.json()
            except Exception:
                err_body = response.text
            print(
                f"  {key_label}: Rate limited (HTTP {response.status_code}). "
                f"Marking exhausted, rotating...",
                file=sys.stderr,
            )
            key_status[key_index] = {"exhausted": True, "error": str(err_body)}
            exhausted_count += 1
            continue

        # Other HTTP errors
        if response.status_code != 200:
            try:
                err_body = response.json()
            except Exception:
                err_body = response.text
            print(f"  {key_label}: HTTP {response.status_code}: {err_body}", file=sys.stderr)
            continue

        # Parse response
        try:
            data = response.json()
        except json.JSONDecodeError:
            print(f"  {key_label}: Invalid JSON response.", file=sys.stderr)
            continue

        raw_text = (
            data.get("choices", [{}])[0]
            .get("message", {})
            .get("content", "")
        )
        if not raw_text:
            print(f"  {key_label}: Empty response content.", file=sys.stderr)
            continue

        cleaned = strip_markdown_fences(raw_text)
        try:
            result = json.loads(cleaned)
        except json.JSONDecodeError:
            json_match = re.search(r"\{[^{}]*\}", cleaned, re.DOTALL)
            if json_match:
                try:
                    result = json.loads(json_match.group())
                except json.JSONDecodeError:
                    print(f"  {key_label}: Could not parse JSON.", file=sys.stderr)
                    continue
            else:
                print(f"  {key_label}: No JSON found in response.", file=sys.stderr)
                continue

        print(f"  {key_label}: Success!", file=sys.stderr)
        return result, None, False

    # All OpenRouter keys exhausted — try Anthropic backup if available
    all_openrouter_exhausted = exhausted_count >= len(api_keys)
    if all_openrouter_exhausted and anthropic_key:
        print("  [Backup] All OpenRouter keys exhausted. Trying Anthropic...", file=sys.stderr)
        try:
            response = _call_anthropic_single(full_prompt, anthropic_key)
            if _is_anthropic_rate_limit(response):
                print(
                    f"  [Backup] Anthropic rate limited (HTTP {response.status_code}).",
                    file=sys.stderr,
                )
                return None, "Anthropic rate limited.", True
            if response.status_code != 200:
                print(
                    f"  [Backup] Anthropic HTTP {response.status_code}: {response.text[:200]}",
                    file=sys.stderr,
                )
                return None, f"Anthropic HTTP {response.status_code}", True
            result = _parse_anthropic_response(response, "[Backup] Anthropic")
            if result:
                print("  [Backup] Anthropic: Success!", file=sys.stderr)
                return result, None, False
            return None, "Anthropic returned no parseable JSON.", True
        except Exception as e:
            print(f"  [Backup] Anthropic error: {e}", file=sys.stderr)
            return None, f"Anthropic error: {e}", True

    return None, "All keys tried, none succeeded.", all_openrouter_exhausted


def save_output(results: list[dict], output_path: str) -> None:
    """Save results to an Excel file."""
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Emails"

    # Write headers
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = openpyxl.styles.Font(bold=True)

    # Write data rows
    for row_idx, result in enumerate(results, start=2):
        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=result.get(col_name, ""))

    # Auto-adjust column widths
    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        max_length = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, min(len(str(cell.value)), 60))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

    wb.save(output_path)
    print(f"\nOutput saved to: {output_path}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Batch email writer for all lead chunks.")
    parser.add_argument(
        "--input",
        default=None,
        help="Path to a specific leads XLSX file (e.g. Quinx/Leads/London_UK_Cafes_Leads.xlsx). "
             "If omitted, reads all leads_chunk_*.xlsx files from the leads/ folder.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output xlsx path. Defaults to Quinx/Emails/{Name}_Emails.xlsx when --input is given, "
             f"or {DEFAULT_OUTPUT} otherwise.",
    )
    parser.add_argument(
        "--start-from",
        type=int,
        default=1,
        help="1-based lead index to start from (skip earlier leads, preserving their existing output rows)",
    )
    parser.add_argument(
        "--campaign",
        default=None,
        help="Campaign config name (e.g. 'quinx_ai') or full path to a campaign JSON file.",
    )
    args = parser.parse_args()

    # Resolve output path
    if args.output is None:
        if args.input:
            # Derive: Quinx/Emails/{Name}_Emails.xlsx from Quinx/Leads/{Name}_Leads.xlsx
            input_filename = os.path.basename(args.input)
            output_filename = input_filename.replace("_Leads.xlsx", "_Emails.xlsx")
            emails_dir = os.path.join(os.path.dirname(BASE_DIR), "Emails")
            args.output = os.path.join(emails_dir, output_filename)
        else:
            args.output = DEFAULT_OUTPUT

    # Load environment
    env_path = os.path.join(BASE_DIR, ".env")
    load_dotenv(env_path)

    api_keys = load_api_keys()
    if not api_keys:
        print("No OpenRouter API keys found in .env.", file=sys.stderr)
        sys.exit(1)

    anthropic_key = load_anthropic_key()
    if anthropic_key:
        print(f"Loaded {len(api_keys)} OpenRouter key(s) + Anthropic backup.", file=sys.stderr)
    else:
        print(f"Loaded {len(api_keys)} API key(s).", file=sys.stderr)

    # Load campaign config
    campaign_config = load_campaign_config(args.campaign)
    if campaign_config:
        print(f"Campaign: {campaign_config.get('campaignName', args.campaign)}", file=sys.stderr)
    else:
        print("No campaign config loaded — service placeholders will be 'Unknown'.", file=sys.stderr)

    # Resolve campaign context and sign-off from config or fall back to defaults
    campaign_context = campaign_config.get('serviceContext', '') or DEFAULT_CAMPAIGN_CONTEXT
    sign_off = campaign_config.get('signOff', '') or DEFAULT_SIGN_OFF

    # Load master prompt template
    template = load_master_prompt()

    # Read leads (from specific file or all chunk files)
    leads = read_all_leads(args.input)
    total = len(leads)
    source = args.input or "all chunk files"
    print(f"Found {total} leads from {source}.\n", file=sys.stderr)

    start_from = max(1, args.start_from)

    # If resuming, load existing rows from the output file to preserve already-done leads
    results = []
    if start_from > 1 and os.path.exists(args.output):
        print(f"Resuming from lead {start_from}. Loading existing rows from {args.output}...", file=sys.stderr)
        wb_existing = openpyxl.load_workbook(args.output)
        ws_existing = wb_existing.active
        existing_headers = [cell.value for cell in ws_existing[1]]
        for row in ws_existing.iter_rows(min_row=2, max_row=start_from, values_only=True):
            row_dict = {}
            for col_name, val in zip(existing_headers, row):
                row_dict[col_name] = val if val is not None else ""
            results.append(row_dict)
        wb_existing.close()
        print(f"  Loaded {len(results)} existing row(s).", file=sys.stderr)

    # Track key status across the batch
    key_status = {}  # {key_index: {"exhausted": bool, "error": str}}

    success_count = 0
    fail_count = 0
    skip_count = 0

    for idx, lead in enumerate(leads, start=1):
        # Skip leads before start_from
        if idx < start_from:
            continue
        biz_name = lead.get("business_name", "Unknown")
        chunk_file = lead.get("_chunk_file", "")
        print(
            f"[{idx}/{total}] Processing: {biz_name} ({chunk_file})",
            file=sys.stderr,
        )

        # Build context (enriches lead with scraped website data)
        context, enrichment_exhausted = build_context(lead, api_keys, key_status, campaign_config)

        if enrichment_exhausted:
            print(
                f"\n{'='*60}\n"
                f"ALL API KEYS EXHAUSTED during enrichment. Stopping batch.\n"
                f"Processed {idx-1}/{total} leads before stopping.\n"
                f"{'='*60}",
                file=sys.stderr,
            )
            results.append({
                "business_name": biz_name,
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "website": lead.get("website", ""),
                "city": lead.get("city", ""),
                "category": lead.get("category", ""),
                "niche": lead.get("niche", ""),
                "subject": "",
                "body": "",
                "status": "api_exhausted",
                "failure_reason": "All API keys exhausted during enrichment",
                "chunk_file": chunk_file,
            })
            for remaining_lead in leads[idx:]:
                results.append({
                    "business_name": remaining_lead.get("business_name", ""),
                    "email": remaining_lead.get("email", ""),
                    "phone": remaining_lead.get("phone", ""),
                    "website": remaining_lead.get("website", ""),
                    "city": remaining_lead.get("city", ""),
                    "category": remaining_lead.get("category", ""),
                    "niche": remaining_lead.get("niche", ""),
                    "subject": "",
                    "body": "",
                    "status": "not_processed",
                    "failure_reason": "Batch stopped — API keys exhausted",
                    "chunk_file": remaining_lead.get("_chunk_file", ""),
                })
            break

        # Skip if missing required fields
        if not context["businessName"] or not context["city"] or not context["category"]:
            print(f"  SKIP: Missing required fields.", file=sys.stderr)
            results.append({
                "business_name": biz_name,
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "website": lead.get("website", ""),
                "city": lead.get("city", ""),
                "category": lead.get("category", ""),
                "niche": lead.get("niche", ""),
                "subject": "",
                "body": "",
                "status": "skipped",
                "failure_reason": "Missing required fields",
                "chunk_file": chunk_file,
            })
            skip_count += 1
            continue

        # Inject campaign-level fields into context
        context["campaignContext"] = campaign_context
        context["signOff"] = sign_off

        # Build prompt
        prompt = substitute_placeholders(template, context)

        # First attempt
        result, error, all_exhausted = call_api_with_rotation(
            prompt, api_keys, key_status, anthropic_key=anthropic_key
        )

        if all_exhausted:
            print(
                f"\n{'='*60}\n"
                f"ALL API KEYS EXHAUSTED. Stopping batch.\n"
                f"Processed {idx-1}/{total} leads before stopping.\n"
                f"{'='*60}",
                file=sys.stderr,
            )
            # Save whatever we have and mark remaining as skipped
            results.append({
                "business_name": biz_name,
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "website": lead.get("website", ""),
                "city": lead.get("city", ""),
                "category": lead.get("category", ""),
                "niche": lead.get("niche", ""),
                "subject": "",
                "body": "",
                "status": "api_exhausted",
                "failure_reason": "All API keys exhausted",
                "chunk_file": chunk_file,
            })
            # Mark remaining leads
            for remaining_lead in leads[idx:]:
                results.append({
                    "business_name": remaining_lead.get("business_name", ""),
                    "email": remaining_lead.get("email", ""),
                    "phone": remaining_lead.get("phone", ""),
                    "website": remaining_lead.get("website", ""),
                    "city": remaining_lead.get("city", ""),
                    "category": remaining_lead.get("category", ""),
                    "niche": remaining_lead.get("niche", ""),
                    "subject": "",
                    "body": "",
                    "status": "not_processed",
                    "failure_reason": "Batch stopped — API keys exhausted",
                    "chunk_file": remaining_lead.get("_chunk_file", ""),
                })
            break

        # Validate first attempt
        if result:
            validation_error = validate_email_output(result, context)
            if validation_error:
                print(f"  Validation failed: {validation_error}", file=sys.stderr)
                print(f"  Retrying with correction...", file=sys.stderr)

                # Retry with reason
                time.sleep(DELAY_BETWEEN_CALLS)
                result2, error2, all_exhausted2 = call_api_with_rotation(
                    prompt, api_keys, key_status, retry_reason=validation_error,
                    anthropic_key=anthropic_key,
                )

                if all_exhausted2:
                    print(f"\n  ALL API KEYS EXHAUSTED during retry. Stopping.", file=sys.stderr)
                    results.append({
                        "business_name": biz_name,
                        "email": lead.get("email", ""),
                        "phone": lead.get("phone", ""),
                        "website": lead.get("website", ""),
                        "city": lead.get("city", ""),
                        "category": lead.get("category", ""),
                        "niche": lead.get("niche", ""),
                        "subject": "",
                        "body": "",
                        "status": "api_exhausted",
                        "failure_reason": "All API keys exhausted during retry",
                        "chunk_file": chunk_file,
                    })
                    for remaining_lead in leads[idx:]:
                        results.append({
                            "business_name": remaining_lead.get("business_name", ""),
                            "email": remaining_lead.get("email", ""),
                            "phone": remaining_lead.get("phone", ""),
                            "website": remaining_lead.get("website", ""),
                            "city": remaining_lead.get("city", ""),
                            "category": remaining_lead.get("category", ""),
                            "niche": remaining_lead.get("niche", ""),
                            "subject": "",
                            "body": "",
                            "status": "not_processed",
                            "failure_reason": "Batch stopped — API keys exhausted",
                            "chunk_file": remaining_lead.get("_chunk_file", ""),
                        })
                    break

                if result2:
                    validation_error2 = validate_email_output(result2, context)
                    if validation_error2:
                        # Second attempt also failed
                        print(f"  Retry also failed: {validation_error2}", file=sys.stderr)
                        results.append({
                            "business_name": biz_name,
                            "email": lead.get("email", ""),
                            "phone": lead.get("phone", ""),
                            "website": lead.get("website", ""),
                            "city": lead.get("city", ""),
                            "category": lead.get("category", ""),
                            "niche": lead.get("niche", ""),
                            "subject": result2.get("subject", ""),
                            "body": result2.get("body", ""),
                            "status": "email_write_failed",
                            "failure_reason": validation_error2,
                            "chunk_file": chunk_file,
                        })
                        fail_count += 1
                    else:
                        # Retry succeeded
                        print(f"  Retry succeeded!", file=sys.stderr)
                        results.append({
                            "business_name": biz_name,
                            "email": lead.get("email", ""),
                            "phone": lead.get("phone", ""),
                            "website": lead.get("website", ""),
                            "city": lead.get("city", ""),
                            "category": lead.get("category", ""),
                            "niche": lead.get("niche", ""),
                            "subject": result2["subject"],
                            "body": result2["body"],
                            "status": "ready_to_send",
                            "failure_reason": "",
                            "chunk_file": chunk_file,
                        })
                        success_count += 1
                else:
                    results.append({
                        "business_name": biz_name,
                        "email": lead.get("email", ""),
                        "phone": lead.get("phone", ""),
                        "website": lead.get("website", ""),
                        "city": lead.get("city", ""),
                        "category": lead.get("category", ""),
                        "niche": lead.get("niche", ""),
                        "subject": "",
                        "body": "",
                        "status": "email_write_failed",
                        "failure_reason": error2 or "API call failed on retry",
                        "chunk_file": chunk_file,
                    })
                    fail_count += 1
            else:
                # First attempt passed validation
                print(f"  Email generated successfully!", file=sys.stderr)
                results.append({
                    "business_name": biz_name,
                    "email": lead.get("email", ""),
                    "phone": lead.get("phone", ""),
                    "website": lead.get("website", ""),
                    "city": lead.get("city", ""),
                    "category": lead.get("category", ""),
                    "niche": lead.get("niche", ""),
                    "subject": result["subject"],
                    "body": result["body"],
                    "status": "ready_to_send",
                    "failure_reason": "",
                    "chunk_file": chunk_file,
                })
                success_count += 1
        else:
            # First attempt failed
            print(f"  FAILED: {error}", file=sys.stderr)
            results.append({
                "business_name": biz_name,
                "email": lead.get("email", ""),
                "phone": lead.get("phone", ""),
                "website": lead.get("website", ""),
                "city": lead.get("city", ""),
                "category": lead.get("category", ""),
                "niche": lead.get("niche", ""),
                "subject": "",
                "body": "",
                "status": "email_write_failed",
                "failure_reason": error or "Unknown error",
                "chunk_file": chunk_file,
            })
            fail_count += 1

        # Delay between leads
        time.sleep(DELAY_BETWEEN_CALLS)

        # Periodic save every 10 leads
        if idx % 10 == 0:
            save_output(results, args.output)
            print(f"  [Checkpoint saved: {idx}/{total} processed]", file=sys.stderr)


    # Final save
    save_output(results, args.output)

    # Summary
    print(
        f"\n{'='*60}\n"
        f"BATCH COMPLETE\n"
        f"{'='*60}\n"
        f"  Total leads:    {total}\n"
        f"  Successful:     {success_count}\n"
        f"  Failed:         {fail_count}\n"
        f"  Skipped:        {skip_count}\n"
        f"  Not processed:  {total - success_count - fail_count - skip_count}\n"
        f"  Output file:    {args.output}\n"
        f"{'='*60}",
        file=sys.stderr,
    )


if __name__ == "__main__":
    main()
