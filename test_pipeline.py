#!/usr/bin/env python3
"""
Quinx AI — Comprehensive Function Test Suite
=============================================
Tests all major functions across:
  - run_pipeline.py (pipeline orchestrator)
  - Email_Scrap/tools/google_maps_search.py
  - Email_Writer/tools/batch_write_emails.py
  - Email_Writer/tools/write_email.py
  - Email_Scrap SMTP/IMAP connectivity

Run:
    python test_pipeline.py
    python test_pipeline.py --quick     # skip slow API/network tests
"""

import argparse
import csv
import imaplib
import io
import json
import os
import re
import smtplib
import ssl
import sys
import tempfile
import time
import traceback
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Setup paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
EMAIL_SCRAP_DIR  = BASE_DIR / "Email_Scrap"
EMAIL_WRITER_DIR = BASE_DIR / "Email_Writer"
EMAIL_SENDER_DIR = BASE_DIR / "Email_Sender"

sys.path.insert(0, str(BASE_DIR))
sys.path.insert(0, str(EMAIL_WRITER_DIR))

# ---------------------------------------------------------------------------
# Test harness
# ---------------------------------------------------------------------------
PASS = "[PASS]"
FAIL = "[FAIL]"
SKIP = "[SKIP]"

results: list[tuple[str, str, str]] = []   # (test_name, status, detail)


def test(name: str, fn, *args, **kwargs):
    """Run a single test function and record the result."""
    try:
        fn(*args, **kwargs)
        results.append((name, PASS, ""))
        print(f"  {PASS}  {name}")
    except AssertionError as e:
        results.append((name, FAIL, str(e)))
        print(f"  {FAIL}  {name}")
        print(f"         └─ {e}")
    except Exception as e:
        results.append((name, FAIL, f"{type(e).__name__}: {e}"))
        print(f"  {FAIL}  {name}")
        print(f"         └─ {type(e).__name__}: {e}")


def skip(name: str, reason: str = ""):
    results.append((name, SKIP, reason))
    print(f"  {SKIP}  {name}  ({reason})")


def section(title: str):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")


# ---------------------------------------------------------------------------
# 1. run_pipeline.py — pure utility functions
# ---------------------------------------------------------------------------

def test_slugify():
    from run_pipeline import slugify
    assert slugify("New York City USA") == "new-york-city-usa"
    assert slugify("Café & Bakery!") == "caf-bakery"
    assert slugify("Tokyo123") == "tokyo123"
    assert slugify("  spaces  ") == "spaces"


def test_banner_runs():
    from run_pipeline import banner
    # Should not raise
    banner("Test Banner Message")


def test_newest_file_returns_none_for_missing():
    from run_pipeline import newest_file
    result = newest_file("/nonexistent/path/*.xyz")
    assert result is None, f"Expected None got {result}"


def test_newest_file_with_real_files():
    from run_pipeline import newest_file
    import glob
    # Use any .py files we know exist
    result = newest_file(str(BASE_DIR / "*.py"))
    assert result is not None, "Expected to find at least one .py file"
    assert result.exists(), f"Returned path doesn't exist: {result}"


def test_load_smtp_password():
    from run_pipeline import _load_smtp_password
    pwd = _load_smtp_password()
    assert isinstance(pwd, str), "Password must be a string"
    assert len(pwd) > 0, "Password is empty — check Email_Sender/.env"
    print(f"         └─ Password loaded (len={len(pwd)})")


# ---------------------------------------------------------------------------
# 2. convert_to_chunks — core CSV → XLSX bridge
# ---------------------------------------------------------------------------

def test_convert_to_chunks_basic():
    from run_pipeline import convert_to_chunks

    # Build a small temporary CSV
    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv",
                                     delete=False, encoding="utf-8",
                                     newline="") as f:
        csv_path = Path(f.name)
        writer = csv.DictWriter(f, fieldnames=[
            "business_name", "email", "phone", "website", "city",
            "category", "owner_name"
        ])
        writer.writeheader()
        writer.writerows([
            {"business_name": "Joe's Cafe", "email": "joe@cafe.com",
             "phone": "123", "website": "cafe.com", "city": "London",
             "category": "cafe", "owner_name": "Joe"},
            {"business_name": "Sam's Bakery", "email": "sam@bakery.com",
             "phone": "456", "website": "bakery.com", "city": "Paris",
             "category": "bakery", "owner_name": ""},
            # Row without email — should be skipped
            {"business_name": "No Email Place", "email": "",
             "phone": "789", "website": "", "city": "NYC",
             "category": "misc", "owner_name": ""},
        ])

    try:
        count = convert_to_chunks(csv_path, niche="cafes", chunk_size=5)
        assert count == 2, f"Expected 2 leads (email-only rows), got {count}"

        # Verify chunk file was created
        leads_dir = EMAIL_WRITER_DIR / "leads"
        chunks = list(leads_dir.glob("leads_chunk_*.xlsx"))
        assert len(chunks) >= 1, "No chunk XLSX files created"

        # Verify owner_name fallback
        wb = openpyxl.load_workbook(chunks[0])
        ws = wb.active
        headers = [c.value for c in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        owner_col = headers.index("owner_name")
        for row in rows:
            biz = row[headers.index("business_name")]
            owner = row[owner_col]
            if biz == "Sam's Bakery":
                assert "Sam's Bakery" in owner, f"Fallback owner wrong: {owner}"
    finally:
        csv_path.unlink(missing_ok=True)


def test_convert_to_chunks_with_max_leads():
    from run_pipeline import convert_to_chunks

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv",
                                     delete=False, encoding="utf-8",
                                     newline="") as f:
        csv_path = Path(f.name)
        writer = csv.DictWriter(f, fieldnames=[
            "business_name", "email", "phone", "website", "city",
            "category", "owner_name"
        ])
        writer.writeheader()
        for i in range(10):
            writer.writerow({
                "business_name": f"Biz {i}", "email": f"biz{i}@example.com",
                "phone": "", "website": "", "city": "NYC",
                "category": "test", "owner_name": ""
            })

    try:
        count = convert_to_chunks(csv_path, niche="test", chunk_size=5,
                                   max_leads=3)
        assert count == 3, f"Expected 3 leads after cap, got {count}"
    finally:
        csv_path.unlink(missing_ok=True)


def test_convert_to_chunks_chunk_splitting():
    from run_pipeline import convert_to_chunks

    with tempfile.NamedTemporaryFile(mode="w", suffix=".csv",
                                     delete=False, encoding="utf-8",
                                     newline="") as f:
        csv_path = Path(f.name)
        writer = csv.DictWriter(f, fieldnames=[
            "business_name", "email", "phone", "website", "city",
            "category", "owner_name"
        ])
        writer.writeheader()
        for i in range(7):
            writer.writerow({
                "business_name": f"Biz {i}", "email": f"biz{i}@test.com",
                "phone": "", "website": "", "city": "Tokyo",
                "category": "restaurant", "owner_name": f"Owner {i}"
            })

    leads_dir = EMAIL_WRITER_DIR / "leads"
    # Remove old chunks first
    for f in leads_dir.glob("leads_chunk_*.xlsx"):
        f.unlink()

    try:
        count = convert_to_chunks(csv_path, niche="restaurants", chunk_size=3)
        assert count == 7, f"Expected 7 leads, got {count}"

        chunks = sorted(leads_dir.glob("leads_chunk_*.xlsx"))
        assert len(chunks) == 3, f"Expected 3 chunks (7 rows, size=3), got {len(chunks)}"

        # Chunk 1: 3, Chunk 2: 3, Chunk 3: 1
        wb1 = openpyxl.load_workbook(chunks[0])
        rows1 = list(wb1.active.iter_rows(min_row=2, values_only=True))
        assert len(rows1) == 3, f"Chunk 1 should have 3 rows, got {len(rows1)}"

        wb3 = openpyxl.load_workbook(chunks[2])
        rows3 = list(wb3.active.iter_rows(min_row=2, values_only=True))
        assert len(rows3) == 1, f"Chunk 3 should have 1 row, got {len(rows3)}"
    finally:
        csv_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# 3. send_emails — validation & SMTP/IMAP connectivity (non-destructive)
# ---------------------------------------------------------------------------

def test_smtp_connectivity():
    """Test that SMTP login works without sending any email."""
    from run_pipeline import _load_smtp_password, SMTP_HOST, SMTP_PORT, FROM_EMAIL
    password = _load_smtp_password()
    assert len(password) > 0, "No SMTP password found"

    ctx = ssl.create_default_context()
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=ctx) as smtp:
        smtp.login(FROM_EMAIL, password)
    # If we get here, login succeeded
    print(f"         └─ SMTP login OK ({FROM_EMAIL})")


def test_imap_connectivity():
    """Test IMAP login and Sent folder detection."""
    from run_pipeline import (_load_smtp_password, IMAP_HOST, IMAP_PORT,
                               FROM_EMAIL, _find_sent_folder)
    password = _load_smtp_password()
    assert len(password) > 0, "No IMAP password found"

    with imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT) as imap:
        imap.login(FROM_EMAIL, password)
        sent_folder = _find_sent_folder(imap)
        assert sent_folder, "Could not detect Sent folder"
        print(f"         └─ IMAP login OK, Sent folder='{sent_folder}'")


def test_find_sent_folder_fallback():
    """Unit test _find_sent_folder with a mock IMAP object."""
    from run_pipeline import _find_sent_folder

    class MockIMAP:
        def list(self):
            return "OK", [b'(\\HasNoChildren) "." "Sent"',
                          b'(\\HasNoChildren) "." "INBOX"']
        def select(self, folder):
            return "OK", [b"1"]
        def close(self):
            pass

    folder = _find_sent_folder(MockIMAP())
    assert folder == "Sent", f"Expected 'Sent', got '{folder}'"


def test_send_emails_missing_password(tmp_path):
    """send_emails should exit(1) if password is missing."""
    from run_pipeline import send_emails
    # Create a fake Excel with valid data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["business_name", "email", "subject", "body"])
    ws.append(["Test Co", "test@example.com", "Hello", "Dear Test Co..."])
    xlsx = tmp_path / "fake_output.xlsx"
    wb.save(xlsx)

    # Patch _load_smtp_password to return empty string
    import run_pipeline as rp
    original = rp._load_smtp_password
    rp._load_smtp_password = lambda: ""
    try:
        exited = False
        try:
            send_emails(xlsx)
        except SystemExit as e:
            exited = True
            assert e.code != 0, "Expected non-zero exit code"
        assert exited, "send_emails should have called sys.exit() with empty password"
    finally:
        rp._load_smtp_password = original


# ---------------------------------------------------------------------------
# 4. Email_Writer — write_email.py utilities
# ---------------------------------------------------------------------------

def test_write_email_imports():
    """All key functions from write_email.py should import correctly."""
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.write_email import (
        load_api_keys,
        load_anthropic_key,
        load_master_prompt,
        substitute_placeholders,
        strip_markdown_fences,
        count_words,
    )
    print("         └─ All write_email imports OK")


def test_count_words():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.write_email import count_words
    assert count_words("hello world") == 2
    assert count_words("  one  ") == 1
    assert count_words("") == 0
    assert count_words("a b c d e") == 5


def test_strip_markdown_fences():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.write_email import strip_markdown_fences
    raw = '```json\n{"key": "value"}\n```'
    cleaned = strip_markdown_fences(raw)
    assert "```" not in cleaned
    parsed = json.loads(cleaned)
    assert parsed["key"] == "value"


def test_substitute_placeholders():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.write_email import substitute_placeholders
    template = "Hello {{ownerName}}, I noticed {{businessName}} in {{city}}."
    context = {
        "ownerName": "Joe",
        "businessName": "Joe's Cafe",
        "city": "London",
    }
    result = substitute_placeholders(template, context)
    assert "Joe" in result
    assert "Joe's Cafe" in result
    assert "London" in result
    assert "{{" not in result, "Unreplaced placeholders remain"


def test_load_master_prompt():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.write_email import load_master_prompt
    template = load_master_prompt()
    assert isinstance(template, str), "Template should be a string"
    assert len(template) > 100, "Template too short — likely not loading correctly"
    print(f"         └─ Master prompt loaded ({len(template)} chars)")


def test_load_api_keys():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from dotenv import load_dotenv
    load_dotenv(str(EMAIL_WRITER_DIR / ".env"))
    from tools.write_email import load_api_keys
    keys = load_api_keys()
    assert isinstance(keys, list), "load_api_keys should return a list"
    assert len(keys) > 0, "No API keys found — check Email_Writer/.env"
    print(f"         └─ Loaded {len(keys)} OpenRouter key(s)")


# ---------------------------------------------------------------------------
# 5. Email_Writer — batch_write_emails.py utilities
# ---------------------------------------------------------------------------

def test_validate_email_output_valid():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import validate_email_output
    # Build a body with 90-130 words that contains the business name
    biz = "Joe's Cafe"
    # Each sentence = ~23 words, × 4 = 92 core words + greeting/sign-off = ~100 total
    filler = ("We help businesses like yours grow repeat customers through AI-powered loyalty programs and campaigns. "
              "Our platform is trusted by over 200 local restaurants and cafes. ") * 4
    body = f"Hi Joe,\n\nI came across {biz} and was really impressed by what you've built. {filler}Would love to connect and share more.\n\nBest,\nSahil"
    result = {
        "subject": "Quick question for you",
        "body": body,
    }
    context = {"businessName": biz}
    error = validate_email_output(result, context)
    assert error is None, f"Expected valid but got error: {error}  (word count: {len(body.split())})"


def test_validate_email_output_subject_too_long():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import validate_email_output
    result = {
        "subject": "This subject line is way too long with many many extra words",
        "body": "Hello Joe's Cafe, " + " word" * 50,
    }
    context = {"businessName": "Joe's Cafe"}
    error = validate_email_output(result, context)
    assert error is not None, "Expected subject-too-long error"
    assert "Subject" in error or "subject" in error


def test_validate_email_output_body_too_short():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import validate_email_output
    # Only 5 words in body — well below MIN_WORDS=90
    result = {
        "subject": "Quick note",
        "body": "Hi Joe's Cafe, call us.",
    }
    context = {"businessName": "Joe's Cafe"}
    error = validate_email_output(result, context)
    assert error is not None, "Expected body-too-short error"
    assert "word" in error.lower() or "between" in error.lower(), f"Unexpected error: {error}"


def test_validate_email_output_missing_biz_name():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import validate_email_output
    # Body is long enough (90+ words) but doesn't mention the business name
    # "We help businesses grow..." = 11 words × 9 = 99 words + "Hello there," = 101 total
    filler = ("We help businesses grow their customer base using AI loyalty programs. ") * 9
    result = {
        "subject": "Quick note",
        "body": "Hello there, " + filler,
    }
    context = {"businessName": "Joe's Cafe"}
    error = validate_email_output(result, context)
    assert error is not None, "Expected business name not found error"
    assert "Business name" in error or "business name" in error.lower() or "does not appear" in error, f"Unexpected error: {error}"


def test_validate_email_output_missing_keys():
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import validate_email_output
    result = {"subject": "Only subject"}
    context = {"businessName": ""}
    error = validate_email_output(result, context)
    assert error is not None, "Expected missing 'body' key error"


def test_save_output(tmp_path):
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import save_output, OUTPUT_COLUMNS
    rows = [
        {col: f"val_{col}_{i}" for col in OUTPUT_COLUMNS}
        for i in range(3)
    ]
    out = str(tmp_path / "test_output.xlsx")
    save_output(rows, out)

    assert os.path.exists(out), "Output Excel file not created"
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == OUTPUT_COLUMNS, f"Headers mismatch: {headers}"
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 3, f"Expected 3 data rows, got {len(data_rows)}"


def test_read_all_leads(tmp_path):
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from tools.batch_write_emails import read_all_leads

    # Create a temp XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    cols = ["business_name", "email", "city", "category", "niche", "owner_name",
            "phone", "website"]
    ws.append(cols)
    ws.append(["Test Biz", "test@biz.com", "NYC", "cafe", "cafes", "Owner", "555", "biz.com"])
    ws.append(["Biz 2", "b2@biz.com", "LA", "bar", "bars", "", "666", ""])
    xl_path = str(tmp_path / "leads_chunk_001.xlsx")
    wb.save(xl_path)

    leads = read_all_leads(xl_path)
    assert len(leads) == 2, f"Expected 2 leads, got {len(leads)}"
    assert leads[0]["business_name"] == "Test Biz"
    assert leads[0]["_chunk_file"] == "leads_chunk_001.xlsx"


# ---------------------------------------------------------------------------
# 6. Email_Scrap — google_maps_search.py utilities
# ---------------------------------------------------------------------------

def test_maps_slugify():
    sys.path.insert(0, str(EMAIL_SCRAP_DIR / "tools"))
    from google_maps_search import slugify as maps_slugify
    assert maps_slugify("New York City USA") == "new-york-city-usa"
    # Non-ASCII stripped — just verify it runs and returns something
    result = maps_slugify("café & boulangerie")
    assert isinstance(result, str) and len(result) > 0
    print(f"         └─ Non-ASCII slug: '{result}'")


# ---------------------------------------------------------------------------
# 7. Google Maps API — live call (requires GEMINI/MAPS key)
# ---------------------------------------------------------------------------

def test_google_maps_api_live():
    """Do a single real Google Maps Places API request (1 place, no file write)."""
    from dotenv import load_dotenv
    load_dotenv(str(EMAIL_SCRAP_DIR / ".env"))
    api_key = os.getenv("GOOGLE_MAPS_API_KEY", "")
    assert api_key, "GOOGLE_MAPS_API_KEY not set in Email_Scrap/.env"

    import requests
    PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": "places.id,places.displayName",
    }
    body = {"textQuery": "cafes in London UK", "pageSize": 1}
    resp = requests.post(PLACES_SEARCH_URL, headers=headers, json=body, timeout=15)
    assert resp.status_code == 200, f"Google Maps API returned HTTP {resp.status_code}: {resp.text[:200]}"
    data = resp.json()
    places = data.get("places", [])
    assert len(places) >= 1, "Google Maps API returned 0 places"
    print(f"         └─ Got {len(places)} place(s). First: {places[0].get('displayName', {}).get('text', '?')}")


# ---------------------------------------------------------------------------
# 8. OpenRouter API — live key test  
# ---------------------------------------------------------------------------

def test_openrouter_api_live():
    """Test that at least one OpenRouter API key is valid."""
    sys.path.insert(0, str(EMAIL_WRITER_DIR))
    from dotenv import load_dotenv
    load_dotenv(str(EMAIL_WRITER_DIR / ".env"))
    from tools.write_email import load_api_keys, _call_openrouter_single, _is_rate_limit_error

    keys = load_api_keys()
    assert keys, "No OpenRouter keys available"

    for i, key in enumerate(keys, 1):
        try:
            resp = _call_openrouter_single(
                'Respond with exactly: {"subject":"ok","body":"ok"}', key
            )
            if _is_rate_limit_error(resp):
                print(f"         └─ Key #{i}: rate limited, trying next...")
                continue
            if resp.status_code == 200:
                print(f"         └─ Key #{i}: HTTP 200 OK")
                return  # At least one key works
        except Exception as e:
            print(f"         └─ Key #{i}: error: {e}")
            continue

    raise AssertionError("All OpenRouter keys returned errors or rate limits")


# ---------------------------------------------------------------------------
# 9. End-to-end: print_finish (no side effects)
# ---------------------------------------------------------------------------

def test_print_finish():
    from run_pipeline import print_finish
    # Use a real path under BASE_DIR so relative_to() works
    fake_output = EMAIL_WRITER_DIR / "emails" / "email_output_10chunks.xlsx"
    if not fake_output.exists():
        fake_output.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        wb.save(fake_output)
    print_finish(42, fake_output)


# ---------------------------------------------------------------------------
# Main runner
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Quinx pipeline test suite")
    parser.add_argument("--quick", action="store_true",
                        help="Skip slow live API/network tests")
    args = parser.parse_args()

    # ── Pure utility tests ────────────────────────────────────────────────
    section("1. run_pipeline.py — Utility Functions")
    test("slugify()",                        test_slugify)
    test("banner() runs without error",       test_banner_runs)
    test("newest_file() → None for missing",  test_newest_file_returns_none_for_missing)
    test("newest_file() → file for *.py",     test_newest_file_with_real_files)
    test("_load_smtp_password()",             test_load_smtp_password)

    # ── CSV → XLSX bridge ─────────────────────────────────────────────────
    section("2. convert_to_chunks() — CSV→XLSX Bridge")
    test("basic 2-lead conversion",           test_convert_to_chunks_basic)
    test("max_leads cap",                     test_convert_to_chunks_with_max_leads)
    test("chunk splitting (7 leads / 3)",     test_convert_to_chunks_chunk_splitting)

    # ── SMTP/IMAP connectivity ────────────────────────────────────────────
    section("3. Email Sending — SMTP & IMAP")
    if args.quick:
        skip("SMTP connectivity",  "quick mode")
        skip("IMAP connectivity",  "quick mode")
    else:
        test("SMTP login (team@tryquinx.com)",    test_smtp_connectivity)
        test("IMAP login + Sent folder detect",    test_imap_connectivity)
    test("_find_sent_folder() mock test",          test_find_sent_folder_fallback)
    test("send_emails() exits on missing password",
         test_send_emails_missing_password, Path(tempfile.mkdtemp()))

    # ── Email_Writer utilities ────────────────────────────────────────────
    section("4. Email_Writer — write_email.py")
    test("imports OK",                        test_write_email_imports)
    test("count_words()",                     test_count_words)
    test("strip_markdown_fences()",           test_strip_markdown_fences)
    test("substitute_placeholders()",         test_substitute_placeholders)
    test("load_master_prompt()",              test_load_master_prompt)
    test("load_api_keys()",                   test_load_api_keys)

    # ── Email validation & batch helpers ─────────────────────────────────
    section("5. Email_Writer — batch_write_emails.py")
    test("validate: valid email passes",       test_validate_email_output_valid)
    test("validate: subject too long",         test_validate_email_output_subject_too_long)
    test("validate: body too short",           test_validate_email_output_body_too_short)
    test("validate: business name missing",    test_validate_email_output_missing_biz_name)
    test("validate: missing keys",             test_validate_email_output_missing_keys)
    test("save_output() creates xlsx",
         test_save_output, Path(tempfile.mkdtemp()))
    test("read_all_leads() from specific xlsx",
         test_read_all_leads, Path(tempfile.mkdtemp()))

    # ── Email_Scrap utilities ─────────────────────────────────────────────
    section("6. Email_Scrap — google_maps_search.py")
    test("slugify()",                          test_maps_slugify)

    # ── Live API tests ────────────────────────────────────────────────────
    section("7. Live API Tests")
    if args.quick:
        skip("Google Maps API (live)",   "quick mode")
        skip("OpenRouter API (live)",    "quick mode")
    else:
        test("Google Maps API — 1 result",        test_google_maps_api_live)
        test("OpenRouter API — at least 1 key OK", test_openrouter_api_live)

    # ── Misc pipeline ─────────────────────────────────────────────────────
    section("8. Misc Pipeline")
    test("print_finish() runs without error",  test_print_finish)

    # ── Summary ───────────────────────────────────────────────────────────
    passed  = sum(1 for _, s, _ in results if s == PASS)
    failed  = sum(1 for _, s, _ in results if s == FAIL)
    skipped = sum(1 for _, s, _ in results if s == SKIP)
    total   = len(results)

    print(f"\n{'='*60}")
    print(f"  TEST SUMMARY")
    print(f"{'='*60}")
    print(f"  Total : {total}")
    print(f"  Passed: {passed}  [OK]")
    print(f"  Failed: {failed}  [!!]")
    print(f"  Skipped:{skipped}  [-]")
    print(f"{'='*60}")

    if failed:
        print("\n  Failed tests:")
        for name, status, detail in results:
            if status == FAIL:
                print(f"    [!!] {name}")
                if detail:
                    print(f"       {detail}")
        sys.exit(1)
    else:
        print("\n  All tests passed! 🎉".encode('ascii', 'replace').decode())


if __name__ == "__main__":
    main()
