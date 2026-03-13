"""
tools/pipeline.py

Full lead generation pipeline wrapper for the Quinx GUI.
Runs all 4 steps for a given niche and list of cities:
  1. google_maps_search.py  -> .tmp/raw_places_*.json
  2. scrape_website_emails.py -> .tmp/enriched_*.json
  3. build_leads_csv.py -> .tmp/leads_*.csv
  4. Convert CSV -> Quinx/Leads/{City}_{Niche}_Leads.xlsx (one file per city)

Usage:
    python tools/pipeline.py --niche "cafes" --cities "London UK,Berlin Germany"
    python tools/pipeline.py --niche "cafes" --cities "London UK" --limit 5

Output:
    Streams progress to stdout; final XLSX files written to Quinx/Leads/
"""

import argparse
import csv
import glob
import os
import re
import subprocess
import sys
import time

# Force UTF-8 stdout so Unicode in child scripts doesn't crash on Windows cp1252
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


def slugify(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def run(cmd: list) -> int:
    """Run a command, streaming output to stdout, return exit code."""
    proc = subprocess.Popen(
        cmd,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        encoding="utf-8",
        errors="replace",
    )
    for line in proc.stdout:
        line = line.rstrip()
        if line:
            print(line, flush=True)
    proc.wait()
    return proc.returncode


def city_to_filename_part(text: str) -> str:
    """Convert 'London UK' or 'Cafes' to a safe filename segment like 'London_UK'."""
    result = re.sub(r"[^\w]", "_", text)
    result = re.sub(r"_+", "_", result)
    return result.strip("_")


def _write_leads_xlsx(rows: list, niche: str, output_path: str) -> None:
    """Write a list of lead dicts to an XLSX file."""
    try:
        import openpyxl
    except ImportError:
        print("[WARN] openpyxl not installed — skipping XLSX write.", flush=True)
        return

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["business_name", "email", "phone", "website", "city", "category", "niche", "owner_name"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        biz_name = row.get("business_name", "").strip()
        ws.cell(row=row_idx, column=1, value=biz_name)
        ws.cell(row=row_idx, column=2, value=row.get("email", "").strip())
        ws.cell(row=row_idx, column=3, value=row.get("phone", "").strip())
        ws.cell(row=row_idx, column=4, value=row.get("website", "").strip())
        ws.cell(row=row_idx, column=5, value=row.get("city", "").strip())
        ws.cell(row=row_idx, column=6, value=row.get("category", "").strip() or niche)
        ws.cell(row=row_idx, column=7, value=niche)
        ws.cell(row=row_idx, column=8, value=row.get("owner_name", "").strip() or f"{biz_name}'s Team")

    wb.save(output_path)


def write_city_leads_xlsx(niche: str, cities: list, niche_slug: str, limit: int, leads_dir: str) -> int:
    """
    Step 4: Read the combined CSV, split rows by city, and write one
    Quinx/Leads/{City}_{Niche}_Leads.xlsx file per city.
    Returns total number of leads written.
    """
    # Find the most-recently-modified CSV for this niche
    pattern = f".tmp/leads_{niche_slug}_*.csv"
    csv_files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    if not csv_files:
        print(f"[WARN] No CSV found matching {pattern}. Skipping XLSX step.", flush=True)
        return 0

    latest_csv = csv_files[0]
    print(f"[XLSX] Reading {latest_csv} ...", flush=True)

    # Read all rows with an email address
    all_rows = []
    with open(latest_csv, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("email", "").strip():
                all_rows.append(row)

    if not all_rows:
        print("[WARN] No rows with emails found in CSV.", flush=True)
        return 0

    os.makedirs(leads_dir, exist_ok=True)
    total_written = 0
    niche_part = city_to_filename_part(niche)

    for city_arg in cities:
        city_part = city_to_filename_part(city_arg)
        # Match CSV rows to this city: check if city arg's first keyword appears
        city_keyword = city_arg.split()[0].lower()
        city_rows = [r for r in all_rows if city_keyword in r.get("city", "").lower()]

        # Fallback: if no match and only one city, use all rows
        if not city_rows:
            if len(cities) == 1:
                city_rows = all_rows
            else:
                print(f"[WARN] No rows matched city '{city_arg}' — skipping.", flush=True)
                continue

        if limit:
            city_rows = city_rows[:limit]

        filename = f"{city_part}_{niche_part}_Leads.xlsx"
        output_path = os.path.join(leads_dir, filename)
        _write_leads_xlsx(city_rows, niche, output_path)
        print(f"[XLSX] {city_arg}: {len(city_rows)} leads → {filename}", flush=True)
        total_written += len(city_rows)

    return total_written


def main():
    parser = argparse.ArgumentParser(description="Quinx full lead generation pipeline")
    parser.add_argument("--niche", required=True, help="Business type, e.g. 'cafes'")
    parser.add_argument("--cities", required=True,
                        help="Comma-separated city list, e.g. 'London UK,Berlin Germany'")
    parser.add_argument("--limit", type=int, default=0,
                        help="Max leads to export to XLSX (0 = no limit)")
    args = parser.parse_args()

    niche = args.niche
    cities = [c.strip() for c in args.cities.split(",") if c.strip()]
    niche_slug = slugify(niche)
    python = sys.executable

    # Resolve paths relative to this script's project root
    script_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Email_Scrap/
    project_root = os.path.dirname(script_dir)                                 # Quinx/
    leads_dir = os.path.join(project_root, "Leads")                            # Quinx/Leads/

    os.makedirs(".tmp", exist_ok=True)

    print(f"[START] Niche: {niche} | Cities: {len(cities)}", flush=True)

    # Step 1: Google Maps search
    print(f"\n-- Step 1 / 4 - Google Maps Search --", flush=True)
    for city in cities:
        city_slug = slugify(city)
        output_path = f".tmp/raw_places_{niche_slug}_{city_slug}.json"

        if os.path.exists(output_path) and os.path.getsize(output_path) > 10:
            print(f"[SKIP] {city} (already cached)", flush=True)
            continue

        print(f"[SEARCH] {city} ...", flush=True)
        rc = run([python, "tools/google_maps_search.py", "--niche", niche, "--city", city])
        if rc != 0:
            print(f"[WARN] Maps search failed for {city} (exit {rc}), skipping.", flush=True)
        time.sleep(1.5)

    # Step 2: Scrape emails
    raw_files = sorted(glob.glob(f".tmp/raw_places_{niche_slug}_*.json"))
    print(f"\n-- Step 2 / 4 - Website Scraping ({len(raw_files)} files) --", flush=True)

    for raw_file in raw_files:
        basename = os.path.basename(raw_file)
        city_slug = basename.replace(f"raw_places_{niche_slug}_", "").replace(".json", "")
        enriched_path = f".tmp/enriched_{niche_slug}_{city_slug}.json"

        if os.path.exists(enriched_path) and os.path.getsize(enriched_path) > 10:
            print(f"[SKIP] {city_slug} (already enriched)", flush=True)
            continue

        print(f"[SCRAPE] {city_slug} ...", flush=True)
        rc = run([python, "tools/scrape_website_emails.py", "--input", raw_file])
        if rc != 0:
            print(f"[WARN] Scraping failed for {city_slug} (exit {rc}).", flush=True)

    # Step 3: Build CSV
    print(f"\n-- Step 3 / 4 - Building CSV --", flush=True)
    rc = run([python, "tools/build_leads_csv.py", "--niche", niche])
    if rc != 0:
        print(f"[ERROR] CSV build failed (exit {rc}). Check logs above.", flush=True)
        sys.exit(1)

    # Step 4: Write per-city XLSX files to Quinx/Leads/
    print(f"\n-- Step 4 / 4 - Exporting to Leads folder --", flush=True)
    count = write_city_leads_xlsx(niche, cities, niche_slug, args.limit, leads_dir)
    if count > 0:
        print(f"[DONE] {count} leads written to {leads_dir}", flush=True)

    # Step 5: Persist freshly-scraped leads into leads.db
    # Wrapped in try/except so DB failures never break the pipeline.
    try:
        sys.path.insert(0, os.path.join(script_dir, "tools"))
        from db_handler import init_db, insert_leads

        # Find the most-recent CSV for this niche (just written by Step 3)
        csv_pattern = f".tmp/leads_{niche_slug}_*.csv"
        csv_files = sorted(glob.glob(csv_pattern), key=os.path.getmtime, reverse=True)
        if csv_files:
            latest_csv = csv_files[0]
            leads_to_insert = []
            with open(latest_csv, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if not row.get("email", "").strip():
                        continue
                    # Derive country from city string (e.g. "London UK" -> "UK")
                    city_str = row.get("city", "").strip()
                    city_parts = city_str.split()
                    country = city_parts[-1].upper() if len(city_parts) > 1 else ""
                    leads_to_insert.append({
                        **row,
                        "country":     country,
                        "niche":       niche,
                        "source_file": os.path.basename(latest_csv),
                        "source":      "google_maps",
                    })

            if leads_to_insert:
                init_db()
                result = insert_leads(leads_to_insert)
                print(f"[DB] Persisted to leads.db — inserted: {result['inserted']}, skipped: {result['skipped']}", flush=True)
    except Exception as e:
        print(f"[WARN] DB persist skipped: {e}", flush=True)

    print(f"\n[DONE] Pipeline complete.", flush=True)


if __name__ == "__main__":
    main()
